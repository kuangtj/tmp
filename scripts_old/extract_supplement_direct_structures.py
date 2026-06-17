#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import re
import json
import uuid
import argparse
from pathlib import Path
from typing import Any, Dict, List, Iterable, Tuple

from rdkit import Chem
from openpyxl import load_workbook

from ipm_eagle.db.sqlite import get_conn


DEFAULT_TASK_TYPES = [
    "supplementary_structure_table",
     "supplementary_assay_table",
]


NAME_KEYS = {
    "compound",
    "compound name",
    "compound_name",
    "compound id",
    "compound_id",
    "cpd",
    "cpd id",
    "id",
    "name",
    "no",
    "no.",
    "entry",
}

SMILES_KEY_PATTERNS = [
    "smiles",
    "canonical_smiles",
    "isomeric_smiles",
]

INCHI_KEY_PATTERNS = [
    "inchi",
]


def uid(prefix: str, *parts: Any) -> str:
    return prefix + "_" + uuid.uuid5(
        uuid.NAMESPACE_URL,
        "|".join(str(x) for x in parts),
    ).hex[:16]


def clean(x: Any) -> str:
    return re.sub(r"\s+", " ", str(x or "").strip())


def jdump(x: Any) -> str:
    return json.dumps(x, ensure_ascii=False, default=str)


def table_exists(conn, table: str) -> bool:
    return conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table,),
    ).fetchone() is not None


def table_cols(conn, table: str) -> set:
    if not table_exists(conn, table):
        return set()
    return {
        r["name"]
        for r in conn.execute(f"PRAGMA table_info({table})").fetchall()
    }


def ensure_col(conn, table: str, col: str, typ: str) -> None:
    if col not in table_cols(conn, table):
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {typ}")


def ensure_tables(conn) -> None:
    conn.execute("""
    CREATE TABLE IF NOT EXISTS stg_structure_candidate (
        candidate_id TEXT PRIMARY KEY,
        doc_id TEXT,
        asset_id TEXT,
        image_path TEXT,
        image_name TEXT,
        candidate_index INTEGER,
        smiles TEXT,
        canonical_smiles TEXT,
        rdkit_valid INTEGER,
        rdkit_error TEXT,
        smiles_source TEXT,
        raw_json TEXT,
        status TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        source_tool TEXT,
        raw_output TEXT,
        molecule_label TEXT,
        bbox_json TEXT,
        raw_context_json TEXT
    )
    """)

    for col, typ in {
        "image_name": "TEXT",
        "smiles_source": "TEXT",
        "raw_json": "TEXT",
        "source_tool": "TEXT",
        "raw_output": "TEXT",
        "molecule_label": "TEXT",
        "bbox_json": "TEXT",
        "raw_context_json": "TEXT",
    }.items():
        ensure_col(conn, "stg_structure_candidate", col, typ)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS stg_component_relation (
        relation_id TEXT PRIMARY KEY,
        doc_id TEXT,
        asset_id TEXT,
        candidate_id TEXT,
        compound_name TEXT,
        component_role TEXT,
        relation_type TEXT,
        evidence_text TEXT,
        figure_ref TEXT,
        confidence REAL,
        review_required INTEGER,
        raw_output TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    conn.commit()


def canonicalize_structure(value: str, source_kind: str) -> Tuple[bool, str, str]:
    value = clean(value)

    if not value:
        return False, "", "empty_structure"

    try:
        if source_kind == "inchi":
            mol = Chem.MolFromInchi(value)
        elif source_kind == "molblock":
            mol = Chem.MolFromMolBlock(value, sanitize=True)
        else:
            mol = Chem.MolFromSmiles(value)

        if mol is None:
            return False, "", "rdkit_mol_none"

        can = Chem.MolToSmiles(mol, canonical=True)

        if "*" in can:
            return False, "", "contains_dummy_atom"

        if re.search(r"\[[^\]]*R[^\]]*\]", can):
            return False, "", "contains_r_group"

        if re.search(r"(^|[^A-Za-z])R\d*([^A-Za-z]|$)", can):
            return False, "", "contains_r_group"

        return True, can, ""

    except Exception as e:
        return False, "", str(e)


def normalize_key(k: Any) -> str:
    return clean(k).lower().replace("-", "_")


def extract_rows_from_table_json(table_json: str) -> List[Dict[str, Any]]:
    if not table_json:
        return []

    try:
        obj = json.loads(table_json)
    except Exception:
        return []

    if isinstance(obj, list):
        if all(isinstance(x, dict) for x in obj):
            return obj

        if obj and isinstance(obj[0], list):
            header = [clean(x) for x in obj[0]]
            rows = []
            for row in obj[1:]:
                if isinstance(row, list):
                    rows.append({
                        header[i] if i < len(header) and header[i] else f"col_{i}": row[i]
                        for i in range(len(row))
                    })
            return rows

    if isinstance(obj, dict):
        for k in ["rows", "data", "table", "items"]:
            v = obj.get(k)
            if isinstance(v, list):
                if all(isinstance(x, dict) for x in v):
                    return v
                if v and isinstance(v[0], list):
                    header = [clean(x) for x in v[0]]
                    rows = []
                    for row in v[1:]:
                        if isinstance(row, list):
                            rows.append({
                                header[i] if i < len(header) and header[i] else f"col_{i}": row[i]
                                for i in range(len(row))
                            })
                    return rows

    return []


def load_csv_rows(path: Path) -> List[Dict[str, Any]]:
    sep = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        return list(csv.DictReader(f, delimiter=sep))


def load_xlsx_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    out = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [clean(x) for x in rows[0]]
        for row in rows[1:]:
            d = {
                header[i] if i < len(header) and header[i] else f"col_{i}": row[i]
                for i in range(len(row))
            }
            d["_sheet"] = ws.title
            out.append(d)

    return out


def load_smi_rows(path: Path) -> List[Dict[str, Any]]:
    out = []

    for i, line in enumerate(path.read_text(encoding="utf-8", errors="ignore").splitlines()):
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        parts = line.split()
        if not parts:
            continue

        smi = parts[0]
        name = parts[1] if len(parts) > 1 else f"compound_{i+1}"

        out.append({
            "compound": name,
            "smiles": smi,
            "_line_no": i + 1,
        })

    return out


def load_sdf_rows(path: Path) -> List[Dict[str, Any]]:
    out = []
    suppl = Chem.SDMolSupplier(str(path), sanitize=True)

    for i, mol in enumerate(suppl):
        if mol is None:
            continue

        name = mol.GetProp("_Name") if mol.HasProp("_Name") else f"compound_{i+1}"
        smi = Chem.MolToSmiles(mol, canonical=True)

        props = {
            k: mol.GetProp(k)
            for k in mol.GetPropNames()
        }

        props.update({
            "compound": name,
            "smiles": smi,
            "_sdf_index": i,
        })

        out.append(props)

    return out


def load_rows_from_file(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []

    suffix = path.suffix.lower()

    if suffix in {".csv", ".tsv", ".tab"}:
        return load_csv_rows(path)

    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx_rows(path)

    if suffix in {".smi", ".smiles"}:
        return load_smi_rows(path)

    if suffix == ".sdf":
        return load_sdf_rows(path)

    return []


def find_compound_name(row: Dict[str, Any]) -> str:
    norm = {normalize_key(k): k for k in row.keys()}

    for k in NAME_KEYS:
        nk = normalize_key(k)
        if nk in norm:
            v = clean(row.get(norm[nk]))
            if v:
                return v

    # fallback: first short non-structure value
    for k, v in row.items():
        nk = normalize_key(k)
        if any(x in nk for x in ["smiles", "inchi", "molfile", "molblock", "structure"]):
            continue

        v = clean(v)
        if v and len(v) <= 80:
            return v

    return ""


def find_structure_value(row: Dict[str, Any]) -> Tuple[str, str, str]:
    for k, v in row.items():
        nk = normalize_key(k)
        val = clean(v)

        if not val:
            continue

        if any(p in nk for p in SMILES_KEY_PATTERNS):
            return val, "smiles", str(k)

        if any(p in nk for p in INCHI_KEY_PATTERNS):
            return val, "inchi", str(k)

        if "molblock" in nk or "molfile" in nk:
            return str(v), "molblock", str(k)

    return "", "", ""


def load_supplement_tasks(conn, doc_id: str, task_types: List[str]):
    marks = ",".join(["?"] * len(task_types))
    params = [doc_id] + task_types

    return conn.execute(f"""
    SELECT
        p.task_id,
        p.doc_id,
        p.asset_id,
        p.asset_type AS task_asset_type,
        p.task_type,
        p.priority,
        p.reason,
        a.asset_type,
        a.page_no,
        a.figure_ref,
        a.table_ref,
        a.file_path,
        a.metadata_json
    FROM planned_tasks p
    JOIN raw_asset a
      ON p.asset_id = a.asset_id
    WHERE p.doc_id=?
      AND p.task_type IN ({marks})
      AND p.status='planned'
    ORDER BY
        CASE p.priority
            WHEN 'high' THEN 1
            WHEN 'medium' THEN 2
            ELSE 3
        END,
        a.page_no,
        p.task_id
    """, params).fetchall()


def load_raw_table_rows(conn, doc_id: str, page_no: int, table_ref: str) -> List[Dict[str, Any]]:
    if table_ref:
        tables = conn.execute("""
        SELECT table_id, table_ref, file_path, table_json
        FROM raw_table
        WHERE doc_id=? AND table_ref=?
        """, (doc_id, table_ref)).fetchall()
    else:
        tables = conn.execute("""
        SELECT table_id, table_ref, file_path, table_json
        FROM raw_table
        WHERE doc_id=? AND page_no=?
        """, (doc_id, page_no)).fetchall()

    rows = []

    for t in tables:
        for i, r in enumerate(extract_rows_from_table_json(t["table_json"])):
            d = dict(r)
            d["_source"] = "raw_table"
            d["_table_id"] = t["table_id"]
            d["_table_ref"] = t["table_ref"]
            d["_row_index"] = i
            rows.append(d)

    return rows


def insert_candidate_and_relation(conn, task, row, row_index: int) -> Dict[str, Any]:
    doc_id = task["doc_id"]
    asset_id = task["asset_id"]
    compound_name = find_compound_name(row)
    structure_value, source_kind, source_col = find_structure_value(row)

    if not compound_name or not structure_value:
        return {
            "inserted": False,
            "reason": "missing_compound_or_structure",
        }

    ok, canonical, err = canonicalize_structure(structure_value, source_kind)

    candidate_id = uid(
        "suppstruct",
        doc_id,
        asset_id,
        compound_name,
        canonical or structure_value,
    )

    raw_context = {
        "task_id": task["task_id"],
        "task_type": task["task_type"],
        "asset_id": asset_id,
        "source": row.get("_source", "file"),
        "source_col": source_col,
        "source_kind": source_kind,
        "row_index": row_index,
        "row": row,
        "structure_source_type": "supplement_table_direct",
        "structure_source_priority": 95,
        "is_full_structure": True,
    }

    status = "ok" if ok else "review_invalid_direct_structure"

    conn.execute("""
    INSERT OR REPLACE INTO stg_structure_candidate
    (
        candidate_id, doc_id, asset_id, image_path, image_name,
        candidate_index, smiles, canonical_smiles,
        rdkit_valid, rdkit_error, smiles_source,
        raw_json, source_tool, raw_output,
        molecule_label, bbox_json, raw_context_json, status
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        candidate_id,
        doc_id,
        asset_id,
        clean(task["file_path"]),
        Path(clean(task["file_path"])).name,
        row_index,
        structure_value if source_kind == "smiles" else canonical,
        canonical,
        int(ok),
        err,
        source_kind,
        jdump(raw_context),
        "supplement_table_direct",
        jdump(raw_context),
        compound_name,
        "",
        jdump(raw_context),
        status,
    ))

    relation_id = uid(
        "rel",
        doc_id,
        asset_id,
        candidate_id,
        compound_name,
        "full",
    )

    conn.execute("""
    INSERT OR REPLACE INTO stg_component_relation
    (
        relation_id, doc_id, asset_id, candidate_id,
        compound_name, component_role, relation_type,
        evidence_text, figure_ref, confidence,
        review_required, raw_output
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        relation_id,
        doc_id,
        asset_id,
        candidate_id,
        compound_name,
        "full",
        "exact_table_row",
        f"Supplementary structure table row gives full structure for {compound_name}.",
        clean(task["table_ref"]),
        0.99 if ok else 0.50,
        0 if ok else 1,
        jdump(raw_context),
    ))

    return {
        "inserted": True,
        "compound_name": compound_name,
        "candidate_id": candidate_id,
        "canonical_smiles": canonical,
        "status": status,
        "error": err,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--doc-id", required=True)
    ap.add_argument("--task-types", default=",".join(DEFAULT_TASK_TYPES))
    ap.add_argument("--overwrite", action="store_true")
    args = ap.parse_args()

    task_types = [x.strip() for x in args.task_types.split(",") if x.strip()]

    conn = get_conn()
    ensure_tables(conn)

    tasks = load_supplement_tasks(conn, args.doc_id, task_types)

    if args.overwrite:
        for t in tasks:
            conn.execute(
                "DELETE FROM stg_component_relation WHERE doc_id=? AND asset_id=?",
                (args.doc_id, t["asset_id"]),
            )
            conn.execute(
                "DELETE FROM stg_structure_candidate WHERE doc_id=? AND asset_id=?",
                (args.doc_id, t["asset_id"]),
            )
        conn.commit()

    outputs = []
    stats = {
        "tasks": len(tasks),
        "rows_seen": 0,
        "inserted": 0,
        "valid": 0,
        "review": 0,
        "skipped": 0,
    }

    for t in tasks:
        rows = []

        rows.extend(
            load_raw_table_rows(
                conn,
                doc_id=args.doc_id,
                page_no=t["page_no"],
                table_ref=t["table_ref"],
            )
        )

        file_rows = load_rows_from_file(Path(clean(t["file_path"])))
        for i, r in enumerate(file_rows):
            d = dict(r)
            d["_source"] = "file"
            d["_row_index"] = i
            rows.append(d)

        for i, row in enumerate(rows):
            stats["rows_seen"] += 1
            out = insert_candidate_and_relation(conn, t, row, i)
            outputs.append(out)

            if not out.get("inserted"):
                stats["skipped"] += 1
                continue

            stats["inserted"] += 1
            if out.get("status") == "ok":
                stats["valid"] += 1
            else:
                stats["review"] += 1

        conn.commit()

    out_dir = Path("data/staging") / args.doc_id
    out_dir.mkdir(parents=True, exist_ok=True)

    report = {
        "doc_id": args.doc_id,
        "task_types": task_types,
        "stats": stats,
        "note": (
            "Direct supplementary structures have highest priority before image recognition "
            "and before component reconstruction."
        ),
    }

    (out_dir / "supplement_direct_structure_report.json").write_text(
        jdump(report),
        encoding="utf-8",
    )

    print(json.dumps(report, ensure_ascii=False, indent=2))
    conn.close()


if __name__ == "__main__":
    main()
